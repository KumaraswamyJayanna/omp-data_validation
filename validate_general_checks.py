""" Verify for business level checks"""

import logging
import os
from datetime import datetime

import openpyxl
import pandas as pd
from config import CATEGORY_NAME, REPORTPATH
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Set up logging


class Generalchecks:

    def __init__(self, datafile, lookupfile, category_name) -> None:
        self.df_datafile = pd.read_excel(datafile)
        self.df_lookupfile = pd.read_excel(lookupfile, sheet_name="Generic_validation")
        self.timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        self.category_name = category_name
        self.report_path = REPORTPATH +f"/highlighted_report_{self.category_name + self.timestamp}.xlsx"

    def create_logger(self):
        logger_report = REPORTPATH +f"report_{CATEGORY_NAME}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
        logging.basicConfig(
            filename=logger_report,
            level=logging.INFO,
            format='%(asctime)s:%(levelname)s:%(message)s'
        )

    def get_mandatory_columns(self):
        mandatory_columns = [self.df_lookupfile.loc[self.df_lookupfile['mandatory_columns']=='Y', 'Fields']]
        logging.info(f"BELOW ARE THE MANDATORY COLUMNS FOR THIS CATEGORY : \n {mandatory_columns}")
        return mandatory_columns

    def check_columns_missing(self):
        print(self.df_lookupfile.columns)
        missing_columns = [col for col in self.df_datafile.columns if col not in self.df_lookupfile['Fields'].to_list()
        ]
        extra_columns = [col for col in self.df_lookupfile['Fields'].to_list() if col not in self.df_datafile.columns
        ]
        if not missing_columns and not extra_columns:
            logging.info("NO MISSING COLUMNS AND  NO EXTRA COLUMNS")

        elif not missing_columns and extra_columns:
            logging.error(f"FOUND EXTRA COLUMNS : {extra_columns}")

        elif missing_columns and not extra_columns:
            logging.error(f"FOUND MISSING COLUMNS : {missing_columns}")

        else:
            logging.error(f"FOUND MISSING COLUMNS : {missing_columns}")
            logging.error(f"FOUND EXTRA COLUMNS: {extra_columns}")

        return missing_columns, extra_columns

    def verify_for_all_null_values(self):
        all_nulls_fileds =[]
        #change here if we want to check only for mandatory columns
        for col in self.df_datafile.columns:
            if self.df_datafile[col].isnull().all():
                all_nulls_fileds.append(col)
        logging.info(f'ALL NULL VALUES COLUMNS: {all_nulls_fileds}')
        return all_nulls_fileds

    def mandatory_columns_null_values(self, mandatory_columns:list):
        mandatory_fields_null_values = []
        for field in mandatory_columns:
            if self.df_datafile[field].isnull:
                mandatory_fields_null_values.append(field)
        logging.info(f'MANDATORY FILEDS NULL VALUES : {mandatory_fields_null_values}')
        return mandatory_fields_null_values

    def verify_dtype(self):
        df_dict_type = self.df_lookupfile[['Fields','dtype']].set_index('Fields').to_dict()['dtype']
        dtype_mismatched_columns = []

        try:
            for col in self.df_datafile.columns:
                if col in df_dict_type.keys():
                    if self.df_datafile[col].dtype != df_dict_type[col]:
                        dtype_mismatched_columns.append(col)
                        logging.error(f"Data Type of the column is not matching {col}")
        except KeyError:
            logging.error("Column not found in lookup file")
        finally:
            logging.info(f"Dtype mismatched columns {dtype_mismatched_columns}")
        return dtype_mismatched_columns


class Report(Generalchecks):

    def __init__(self, datafile, lookupfile, category_name) -> None:
        super().__init__(datafile, lookupfile, category_name)

    def create_report_sheet(self):
        with pd.ExcelWriter(self.report_path, engine='openpyxl') as writer:
            self.df_datafile.to_excel(writer, index=False)
        report_path = os.path.abspath(self.report_path)
        logging.info(f'Reports are copied here : {report_path}')
        return report_path


    def highlight_complete_column(self, report, columns:list, color="FFCCCB"):
        workbook = openpyxl.load_workbook(report)
        sheet = workbook.active
        fill_color_column = PatternFill(start_color=color, end_color=color, fill_type='solid')
        header_map = {cell.value: cell.column for cell in sheet[1] if cell.value}

        column_indices=[]
        for col in columns:
            if isinstance(col, int):
                column_indices.append(col)
            elif isinstance(col, str):
                if col.upper() in openpyxl.utils.cell.get_column_letter(header_map.get(col,0)):
                    column_indices.append(openpyxl.utils.column_index_from_string(col))
                elif col in header_map:
                    column_indices.append(header_map[col])
                else:
                    raise ValueError(f"Column {col} not found in headers")
        for column_index in column_indices:
            for row in range(1, sheet.max_row+1):
                cell = sheet.cell(row=row, column=column_index)
                cell.fill = fill_color_column
        workbook.save(report)

    def highlight_cell(self, index_list, color="FFFF00"):

        wb = openpyxl.load_workbook(self.report_path)
        ws = wb.active

        fill_color = PatternFill(
            start_color=color, end_color=color, fill_type="solid"
        )
        # Apply color to the specified cells
        for row_idx, col_idx in index_list:
            cell = ws.cell(row=row_idx + 1, column=col_idx + 1)
            cell.fill = fill_color
        wb.save(self.report_path)


    def verify_missing_values_in_mandatory_fields(self):

        self.df_report = pd.read_excel(self.report)
        workbook = load_workbook(self.report)
        sheet = workbook.active
        mandatory_columns = self.get_mandatory_columns()
        fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type='solid')
        for col in mandatory_columns:
            null_rows = self.report[self.report[col].isnull()]
            if not null_rows.empty:
                for index in null_rows.index:
                    cell = sheet[f'{chr(65 + self.df_report.columns.get_loc(col))}{index+2}']
                    cell.fill = fill
            else:
                print(f"No Null values in {col}")
