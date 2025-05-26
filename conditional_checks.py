import logging

import openpyxl
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# from utils.lookup_data import Lookupdata


class ConditionalChecks:


    def __init__(self, datafile, lookupfile):
        self.df_datafile=pd.read_excel(datafile)
        self.df_conditional_lookup = pd.read_excel(lookupfile)

    def columns_to_lowercase(self):
        self.df_datafile.columns = self.df_datafile.columns.str.lower()
        self.df_conditional_lookup.columns = self.df_conditional_lookup.columns.str.lower()
        self.df_datafile.columns = self.df_datafile.columns.str.replace(r'[^a-zA-Z0-9]', '', regex=True)
        self.df_conditional_lookup.columns = self.df_conditional_lookup.columns.str.replace(r'[^a-zA-Z0-9]', '', regex=True)

    def highlight_and_add_comments(self, ws, row, col, message, color_fill):
        cell = ws.cell(row=row, column=col)
        cell.fill = color_fill
        cell.comment = openpyxl.comments.Comment(message, "Validation Script")

    def verify_original_name_data(self, report):
        """
        This lookup file consists of the columns name need to be verified and the respective values

        """
        # verifying in the report listed columns expected values are matching or not
        wb = load_workbook(report)
        ws = wb.active
        self.df_datafile = self.df_datafile.applymap(lambda x: x.lower() if isinstance(x, str) else x)
        self.df_conditional_lookup = self.df_conditional_lookup.applymap(lambda x: x.lower() if isinstance(x, str) else x)
        # verify is there a null value
        # Add a comments column for the reasons (in the last column)
        comments_column_index = len(self.df_datafile.columns) + 1
        # Define the color fill for highlighting invalid cells and nulls

        invalid_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        null_fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")
        # Validate columns from data file against lookup file
        for i, column in enumerate(self.df_datafile.columns, 1):

            # Modify here if we need to check only for mandatory columns
            if column in self.df_conditional_lookup.columns:
                # print(f'Conditionaal lookup columns :{column}')
                lookup_values = self.df_conditional_lookup[column].dropna().values  # Get non-null values from lookup column
                for j, value in enumerate(self.df_datafile[column], 2):  # Data rows start from 2 (1 is header)
                    updated_value=''
                    if pd.isna(value):  # Check for null values
                        message =f'NULL value found in {column}'
                        self.highlight_and_add_comments(ws, j, i, "Null value found", null_fill)
                        cell_val = ws.cell(row=j, column=comments_column_index, value=message)
                        existing_value = cell_val.value
                        new_value = column
                        if existing_value:
                            updated_value = f"{existing_value} {new_value}"  # You can use a different separator like a comma
                        else:
                            updated_value = new_value
                        cell_val = updated_value
                    elif value not in lookup_values:  # Check if value is not in the lookup values
                        message =f'Data not matching with lookup in {column}'
                        self.highlight_and_add_comments(ws, j, i, "Value not found in lookup", invalid_fill)
                        cell_val = ws.cell(row=j, column=comments_column_index, value=message)
                        existing_value = cell_val.value
                        new_value = column
                        if existing_value:
                            updated_value = f"{existing_value} {new_value}"  # You can use a different separator like a comma
                        else:
                            updated_value = new_value
                        cell_val = updated_value

            else:
                logging.info(f"Column '{column}' is not checked conditional field verification.")
        wb.save(report)

    def verify_for_non_negative(self, report):
        columns = ["quantity", "Total_Price", "totalprice"]
        wb = load_workbook(report)
        ws = wb.active
        fill_for_invalid = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")
        for column in columns:
            print(f"Verifying for non negative columns {column}")
            if column in self.df_datafile.columns:
                col_index = self.df_datafile.columns.get_loc(column)+1

                for row_id, value in enumerate(self.df_datafile[column], start=2):
                    if value<0:
                        self.highlight_and_add_comments(ws, row_id, col_index, "negative_value", fill_for_invalid)
        wb.save(report)


    def supplier_name_lookup(self, report, filename='Supplier_Alias_Name.csv', id="supplierid", id_name="suppliernameoriginal"):
        wb = load_workbook(report)
        ws = wb.active
        self.df_datafile = self.df_datafile.applymap(lambda x: x.lower() if isinstance(x, str) else x)
        self.df_conditional_lookup = self.df_conditional_lookup.applymap(lambda x: x.lower() if isinstance(x, str) else x)
        mismatch_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
        #  suppliar id suppliar names lookup
        supplier_alias = pd.read_csv(f"lookupdata/{filename}")
        supplier_lookup = supplier_alias.groupby('supplier_id')['alternative_name'].apply(list).reset_index()
        res = dict(zip(supplier_lookup['supplier_id'], supplier_lookup['alternative_name']))
        print("Verifying the suppliers mapping values")
        if id and id_name in self.df_datafile.columns:
            print("Columns found for suppliers")
            for idx, (value, name) in enumerate(zip(self.df_datafile[id], self.df_datafile[id_name]), start=2):
                if (value in res.keys()) and (name.lower() in res[value]):
                    logging.info("verified supplierid and suppliername")
                else:
                    logging.info("supplier id and supplier_name not matching")
                    # id_column_value = int(self.df_datafile.columns.get_loc(id))+1
                    id_name_column_value = int(self.df_datafile.columns.get_loc(id_name))+1
                    # ws.cell(row=idx, column=id_column_value).fill = mismatch_fill  # Highlight 'id' column
                    ws.cell(row=idx, column=id_name_column_value).fill = mismatch_fill

        wb.save(report)


    def client_alias_name_verify(self, report, filename='Client_Alias_Name.csv', id="clientid", id_name="clientnameoriginal"):
        wb = load_workbook(report)
        ws = wb.active
        self.df_datafile = self.df_datafile.applymap(lambda x: x.lower() if isinstance(x, str) else x)
        self.df_conditional_lookup = self.df_conditional_lookup.applymap(lambda x: x.lower() if isinstance(x, str) else x)
        mismatch_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
        #  client id client names lookup
        client_alias = pd.read_csv(f"lookupdata/{filename}")
        client_alias_lookup = client_alias.groupby('client_id')['alternative_name'].apply(list).reset_index()

        client_alias_names = dict(zip(client_alias_lookup['client_id'], client_alias_lookup['alternative_name']))

        if id and id_name in self.df_datafile.columns:
            for idx, (value, name) in enumerate(zip(self.df_datafile[id], self.df_datafile[id_name]), start=2):
                if (value in client_alias_names.keys()) and (name.lower() in client_alias_names[value]):
                    logging.info("verified supplierid and suppliername")
                else:
                    logging.info("supplier id and supplier_name not matching")
                    # id_column_value = int(self.df_datafile.columns.get_loc(id))+1
                    id_name_column_value = int(self.df_datafile.columns.get_loc(id_name))+1
                    # ws.cell(row=idx, column=id_column_value).fill = mismatch_fill  # Highlight 'id' column
                    ws.cell(row=idx, column=id_name_column_value).fill = mismatch_fill

        wb.save(report)


    def verify_price_date(self, report):
        wb = load_workbook(report)
        ws = wb.active
        # Price_Date>=2017-01-01
        if 'price_date' in self.df_datafile.columns:
            price_date_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
            price_date_column_index = self.df_datafile.columns.get_loc('price_date') + 1
            for idx, price_date in enumerate(self.df_datafile['price_date'], start=2):
                if pd.to_datetime(price_date) < pd.to_datetime('2017-01-01'):
                    self.highlight_and_add_comments(ws, idx, price_date_column_index, "Price date is before 2017-01-01", price_date_fill)
        wb.save(report)

    def verify_payment_term(self, report):
        wb = load_workbook(report)
        ws = wb.active
        payment_term_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
        if 'payment_term' in self.df_datafile.columns:
            payment_term_column_index = self.df_datafile.columns.get_loc('payment_term') + 1
            for idx, payment_term in enumerate(self.df_datafile['payment_term'], start=2):
                if not pd.isna(payment_term) and not payment_term.lower().startswith('net') or not payment_term[3:].isdigit():
                    self.highlight_and_add_comments(ws, idx, payment_term_column_index, "Invalid payment term", payment_term_fill)
        wb.save(report)

    def verify_level5_field(self, report):
        wb = load_workbook(report)
        ws = wb.active
        level5_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
        level5_columns = [col for col in self.df_conditional_lookup.columns if col.lower() in ["level 5", "level 5 category"]]

        if 'level5' in self.df_datafile.columns and level5_columns:
            level5_column_index = self.df_datafile.columns.get_loc('level5') + 1
            lookup_values = pd.concat([self.df_conditional_lookup[col].dropna() for col in level5_columns]).unique()
            for idx, level5_value in enumerate(self.df_datafile['level5'], start=2):
                if level5_value.lower() not in lookup_values:
                    self.highlight_and_add_comments(ws, idx, level5_column_index, "Level 5 value not found in lookup", level5_fill)
        wb.save(report)