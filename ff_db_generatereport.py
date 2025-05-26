""" Generate a report of flatfile to Db comparision """

import logging
import os
from datetime import datetime

import numpy as np
import openpyxl
import pandas as pd
from dbconfig import DB_HOST, DB_NAME, DB_PASSWORD, DB_USER
from establish_dbconnection import PostgresLogger
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# from dbconfig import flatfilepath, dbfilepath

# Set up logging
logger_report =f"testlog_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
logging.basicConfig(
    filename=logger_report,
    level=logging.INFO,
    format='%(asctime)s:%(levelname)s:%(message)s'
)


class DbFlatfileReport:
    """
    Class to generate
    a report comparing flatfile data to Db data.
    """
    def __init__(self, flatfile_path, db_path):
        """
        Initialize the class with flatfile and Db paths.

        Args:
            flatfile_path (str): Path to the flatfile.
            db_path (str): Path to the Db data.

        """
        self.flatfile_path = flatfile_path
        self.db_path = db_path
        if self.flatfile_path.endswith('.xlsx'):
            self.df_flatfile = pd.read_excel(self.flatfile_path, engine='openpyxl')
        elif self.flatfile_path.endswith('.csv'):
            self.df_flatfile = pd.read_csv(self.flatfile_path)
        else:
            print("ERROR : Check the flatfile extension")
        self.df_db = pd.read_excel(self.db_path, engine='openpyxl')
        self.report_path = os.path.join(
            os.path.dirname(self.flatfile_path),
            f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        self.highlight_report_path = os.path.join(
            os.path.dirname(self.flatfile_path),
            f"highlight_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        self.sheet_names = ['Flatfile', 'DataInFlatfileNotinDb', 'DataInDbnotinFlatfile']

    def fill_null_values(self):
        """
        Fill null values in the DataFrames.
        This function will fill null values in the DataFrames with 'nan'.
        Returns:
            None
        """
        self.df_flatfile.fillna('nan', inplace=True)
        self.df_db.fillna('nan', inplace=True)


    def get_common_columns(self):
        """
        remove the system_datetime column from both DataFrame.
        conver the price_date column to date format.
        Convert column names to lowercase and re-order them to be the same in both DataFrames.
        Compare flatfile data to Db data and generate a report.
        This function will check for common columns in both DataFrames,
        convert column names to lowercase, and re-order them to be the same in both DataFrames.
        Returns:
            dataframe: DataFrame containing the common columns.
        """
        # Drop the columns
        self.df_flatfile.drop(columns=['System_DateTime'], inplace=True, errors='ignore')
        self.df_db.drop(columns=['system_datetime'], inplace=True, errors='ignore')

        # convert the date time column to date format
        if 'Price_Date' in self.df_flatfile.columns:
            self.df_flatfile['Price_Date'] = pd.to_datetime(self.df_flatfile['Price_Date']).dt.date
        if 'price_date' in self.df_db.columns:
            self.df_db['price_date'] = pd.to_datetime(self.df_db['price_date']).dt.date

        # Convert column names to lowercase
        self.df_flatfile.columns = self.df_flatfile.columns.str.lower()
        self.df_db.columns = self.df_db.columns.str.lower()

        # Ensure both DataFrames have the same columns
        common_columns = self.df_flatfile.columns.intersection(self.df_db.columns)
        self.df_flatfile = self.df_flatfile[common_columns]
        self.df_db = self.df_db[common_columns]

        # Check for missing columns in either DataFrame
        missing_in_flatfile = set(self.df_db.columns) - set(self.df_flatfile.columns)
        missing_in_db = set(self.df_flatfile.columns) - set(self.df_db.columns)
        if missing_in_flatfile:
            logging.warning(f"Missing columns in flatfile: {missing_in_flatfile}")
        if missing_in_db:
            logging.warning(f"Missing columns in Db: {missing_in_db}")

        # Re-order columns to be the same in both DataFrames
        self.df_flatfile = self.df_flatfile.reindex(sorted(self.df_flatfile.columns), axis=1)
        self.df_db = self.df_db.reindex(sorted(self.df_db.columns), axis=1)

    def create_report_sheet(self):
        """
        Create a new Excel sheet for the report.
        This function will create a new sheet in the report Excel file
        and add headers to it.
        Returns:
            None
        """
        # Create a new workbook and add a sheet for the report
        workbook = Workbook()
        if 'Sheet' in workbook.sheetnames:
            std = workbook['Sheet']
            workbook.remove(std)

        for sheetname in self.sheet_names:
            workbook.create_sheet(title=sheetname)
        # save the workbook
        workbook.save(self.highlight_report_path)


    def convert_datetime_columns(self):
        """
        Convert datetime columns to 'YYYY-MM-DD' format.
        This function will check the columns in the DataFrames,
        and if a column is of datetime type, it will convert its data to 'YYYY-MM-DD' format.
        Returns:
            df
        """
        for df in [self.df_flatfile, self.df_db]:
            for column in df.columns:
                if pd.api.types.is_datetime64_any_dtype(df[column]):
                    df[column] = df[column].dt.strftime('%Y-%m-%d')
                    logging.info("Converted Date time format to 'YYYY-MM-DD'")

        return self.df_flatfile, self.df_db

    def append_data_to_report_highlight(self, sheetname, data, columns_to_highlight=None):
        """
        Append data to the report sheet and highlight differences.
        This function will append the data to the specified sheet in the report
        and highlight the differences between the flatfile and Db data.
        Args:
            sheetname (str): Name of the sheet to append data to.
            data (list): Data to append to the sheet.
            columns_to_highlight (list): List of columns to highlight differences in.
        Returns:
            None
        """
        # Load the workbook and select the specified sheet
        workbook = openpyxl.load_workbook(self.highlight_report_path)
        if sheetname in workbook.sheetnames:
            sheet = workbook[sheetname]
            sheet.append(data)
        else:
            print(f"Sheet '{sheetname}' does not exist in the workbook.")
            return
        # Append the data to the worksheet
        last_row = sheet.max_row
        fill =PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        if columns_to_highlight:
            for col in columns_to_highlight:
                    cell = sheet.cell(row=last_row, column=col+1)
                    cell.fill = fill
                    logging.info(f"Highlighting cell {cell.coordinate} in {sheetname} with value {cell.value}")
        workbook.save(self.highlight_report_path)

    def generate_key(self):
        """
        Generate a unique key for each row in the DataFrame.
        This function will create a new column 'key' in the DataFrame
        that concatenates the values of all columns in the row.
        Returns:
            dataframe: DataFrame with the new 'key' column.
        """
        # self.convert_datetime_columns()
        key_column_for_pricepoint = ['file_name', 'product_service_sku_name_normalized']

        # sort the dataframe by file_name
        self.df_flatfile.sort_values(by='file_name', inplace=True)
        self.df_db.sort_values(by='file_name', inplace=True)

        for column in key_column_for_pricepoint:
            if column not in self.df_flatfile.columns:
                logging.error(f"Column '{column}' not found in flatfile DataFrame")
                raise ValueError(f"Column '{column}' not found in DataFrame")
            if column not in self.df_db.columns:
                logging.error(f"Column '{column}' not found in Db DataFrame")
                raise ValueError(f"Column '{column}' not found in DataFrame")
        self.df_flatfile['pseudo_key'] = self.df_flatfile[key_column_for_pricepoint].astype(str).agg('-'.join, axis=1)
        self.df_db['pseudo_key'] = self.df_db[key_column_for_pricepoint].astype(str).agg('-'.join, axis=1)

        # put the pseudo_key column at the beginning ansd remove the special characters
        self.df_flatfile['pseudo_key'] = self.df_flatfile['pseudo_key'].str.replace(r'[^a-zA-Z0-9]', '', regex=True)
        self.df_db['pseudo_key'] = self.df_db['pseudo_key'].str.replace(r'[^a-zA-Z0-9]', '', regex=True)
        # Re-order columns to put 'pseudo_key' at the beginning
        self.df_flatfile = self.df_flatfile[['pseudo_key'] + [col for col in self.df_flatfile.columns if col != 'pseudo_key']]
        self.df_db = self.df_db[['pseudo_key'] + [col for col in self.df_db.columns if col != 'pseudo_key']]

        # Check for duplicate keys
        if self.df_flatfile['pseudo_key'].duplicated().any():
            logging.warning("Duplicate keys found in flatfile DataFrame")
        if self.df_db['pseudo_key'].duplicated().any():
            logging.warning("Duplicate keys found in Db DataFrame")
        # Check for missing keys
        if self.df_flatfile['pseudo_key'].isnull().any():
            logging.warning("Missing keys found in flatfile DataFrame")
        if self.df_db['pseudo_key'].isnull().any():
            logging.warning("Missing keys found in Db DataFrame")


    def clean_data(self, data_list:list):
        cleaned_data = []
        for value in data_list:
            if isinstance(value, np.float64):
                cleaned_data.append('nan' if np.isnan(value) else float(value))
            elif isinstance(value, np.int64):
                cleaned_data.append('nan' if np.isnan(value) else int(value))
            elif isinstance(value, str):
                cleaned_data.append(value.strip())
            else:
                cleaned_data.append(value)
        return cleaned_data

    def compare_dataframes_rowwise_based_on_pseudokey(self):

        self.create_report_sheet()
        columns_data = self.df_flatfile.columns.tolist()
        for sheet in self.sheet_names:
            self.append_data_to_report_highlight(sheetname=sheet, data=columns_data, columns_to_highlight=None)
        # Compare the two DataFrames row-wise based on the 'pseudo_key' column

        for i in range(len(self.df_flatfile)):
            flatfile_row_data = self.df_flatfile.iloc[i].tolist()
            flatfile_row_data = self.clean_data(flatfile_row_data)
            key = flatfile_row_data[0]
            db_row_data = self.df_db[self.df_db['pseudo_key'] == str(key)].values.tolist()
            if not db_row_data:

                logging.warning(f"Row {i} with pseudo_key {key} not found in Db DataFrame")
                self.append_data_to_report_highlight(sheetname='DataInFlatfileNotinDb',
                                                     data=flatfile_row_data,
                                                     columns_to_highlight=None)
                continue

            db_row_data = self.df_db[self.df_db['pseudo_key'] == str(key)]
            db_data_index_values = db_row_data.index.values.tolist()
            db_row_data_for_key = db_row_data.values.tolist()
            data_differences = []

            if len(db_row_data_for_key) == 1:
                # Compare the rows and highlight differences
                if flatfile_row_data == db_row_data_for_key[0]:
                    logging.info(f"Row {i} with pseudo_key {key} is same in both DataFrames")
                    logging.info(flatfile_row_data)
                    logging.info(db_row_data_for_key[0])
                    self.append_data_to_report_highlight(sheetname='Flatfile',
                                                         data=flatfile_row_data,
                                                         columns_to_highlight=None)
                else:
                    logging.info(f"Row {i} with pseudo_key {key} is different in both DataFrames")
                    # differences = [index +1 for index, (a,b) in enumerate(zip(flatfile_row_data, db_row_data_for_key[0])) if str(a) != str(b)]
                    differences = [index for index, (a, b) in enumerate(zip(flatfile_row_data, db_row_data_for_key[0]))
                    if str(a).lower() != str(b).lower() and (logging.info(f"Index: {index}, a: {a}, b: {b}"), True)[1]]
                    self.append_data_to_report_highlight(sheetname='Flatfile',
                                                         data=flatfile_row_data,
                                                         columns_to_highlight=differences)
                    logging.info(flatfile_row_data)
                    logging.info(db_row_data_for_key[0])
                    logging.info(f"{key}={differences}")
                    self.df_db.drop(db_row_data.index, inplace=True)
                    logging.info(f"Dropped the row {db_row_data.index} from Db DataFrame")

            else:
                for data in db_row_data_for_key:
                    logging.info(f"Row {i} with pseudo_key {key} is different in both DataFrames")
                    differences = [index for index, (a, b) in enumerate(zip(flatfile_row_data, data))
                    if str(a).lower() != str(b).lower() and (logging.info(f"Index: {index}, a: {a}, b: {b}"), True)[1]]
                    # differences = [index +1 for index, (a,b) in enumerate(zip(flatfile_row_data, data)) if str(a) != str(b)]
                    data_differences.append(differences)
                    logging.info(f"{key}={differences}")
                    logging.info(flatfile_row_data)
                    logging.info(data)

                if not data_differences:
                    logging.info(f"Row {i} with pseudo_key {key} is same in both DataFrames")
                    #logging.info(f"Exact matches found")
                    self.append_data_to_report_highlight(sheetname='Flatfile',
                                                     data=flatfile_row_data,
                                                     columns_to_highlight=None)

                else:
                    logging.info(f"Row {i} with pseudo_key {key} is different in both DataFrames")
                    original_index= dict(zip(db_data_index_values, data_differences))
                    find_minimum_difference = min(data_differences, key=len)
                    for index, difference in original_index.items():
                        if difference == find_minimum_difference:
                            val = index
                    min_length_index = data_differences.index(find_minimum_difference)
                    logging.info(f"Row {i} with pseudo_key {key} is different in both DataFrames")
                    self.append_data_to_report_highlight(sheetname='Flatfile',
                                                         data=flatfile_row_data,
                                                         columns_to_highlight=data_differences[min_length_index])
                    logging.info(f"{key}={find_minimum_difference}")
                    self.df_db.drop(val, inplace=True)
                    logging.info(f"Dropped the row {val} from Db DataFrame")

        db_data_set = self.df_db.values.tolist()
        logging.info(f"Data in Db not in flatfile: {len(db_data_set)} rows data")
        # write in to the report excel in the sheet DatainFlatfileNotinDb
        for db_data in db_data_set:
            self.append_data_to_report_highlight(sheetname='DataInDbnotinFlatfile',
                                                 data=db_data,
                                                 columns_to_highlight=None)

        logging.info(f"Execution Completed")
        logging.info(f"Report generated at {self.report_path}")
        return self.report_path


if __name__ == "__main__":

    postgres_logger = PostgresLogger(DB_HOST, DB_NAME, DB_USER, DB_PASSWORD)
    logging.info(f'Get the consolidated faltfile')
    flatfile_path  = postgres_logger.get_consolidated_flatfile()
    logging.info(f'Get the DB data for the category')
    db_path = postgres_logger.get_category_db_data()
    report_generator = DbFlatfileReport(flatfile_path, db_path)
    logging.info(f"Flatfile path: {report_generator.flatfile_path}")
    logging.info(f"Db path: {report_generator.db_path}")
    logging.info(f"Get the columns with respect to DB table")
    report_generator.get_common_columns()
    logging.info(f"Generate a pseudo key for each row")
    report_generator.generate_key()
    logging.info(f"Compare the flatfile data to Db data")
    report_generator.compare_dataframes_rowwise_based_on_pseudokey()

