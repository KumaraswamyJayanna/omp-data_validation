"""To generate a category level and file level summary reports"""
import os
import os.path
import shutil
from datetime import datetime

import openpyxl
import pandas as pd
from config import CATEGORY_NAME, REPORTPATH
from openpyxl.styles import PatternFill


class File_Report:

    def __init__(self, reportpath) -> None:
        self.reportpath = reportpath
        self.df_report = pd.read_excel(self.reportpath, sheet_name="Pipeline_Comparission_report")
        self.timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        self.summaryreport = f'summary_{self.timestamp}.xlsx'
        self.counts = {}


    def get_file_names(self):
        '''get the filenames from the report'''
        filenames = self.df_report['File_Name'].unique()
        return filenames


    def count_column_highlights_ofreport(self):
        # count the yellow highlights in the report
        wb = openpyxl.load_workbook(self.reportpath)
        sheet =wb["Pipeline_Comparission_report"]
        highlighted_counts={}

        for col in sheet.columns:
            col_name = col[0].value  # Get column letter
            highlighted_count = 0

            # Check each cell in the column
            for cell in col:
                if cell.fill and cell.fill.start_color.index != '00000000':  # Check if fill is not default (No color)
                    highlighted_count += 1

            # Store the count of highlighted cells for this column
            highlighted_counts[col_name] = highlighted_count

        return highlighted_counts

    def count_column_highlights_byfile(self, data_file):
        # get the columns count based on the filename or we may change to any columns
        wb = openpyxl.load_workbook(data_file)
        sheet =wb["Sheet"]
        highlighted_counts={}

        for col in sheet.columns:
            col_name = col[0].value  # Get column letter
            highlighted_count = 0

            # Check each cell in the column
            for cell in col:
                if cell.fill and cell.fill.start_color.index != '00000000':  # Check if fill is not default (No color)
                    highlighted_count += 1

            # Store the count of highlighted cells for this column
            highlighted_counts[col_name] = highlighted_count

        return highlighted_counts

    def get_columns(self):
        #get the columns of the report
        return self.df_report.columns

    def get_length_by_filename(self):
        # get the count of the data in each files
        data_counts_by_filename = {}
        columnname= "File_Name"
        filenames= self.get_file_names()
        for file in filenames:
            filtered_df = self.df_report[self.df_report[columnname] == file]
            count = filtered_df.shape[0]
            data_counts_by_filename[file]=count
        return data_counts_by_filename


    def filter_by_category(self, column_name, filter_value, output_file_path):
        # filter the values by category
        wb = openpyxl.load_workbook(self.reportpath)
        sheet =wb["Pipeline_Comparission_report"]
        filtered_data =[]

        # Get the column index based on the header
        header_row = sheet[1]  # Assuming the first row contains headers
        column_index = None
        for idx, cell in enumerate(header_row, 1):
            if cell.value == column_name:
                column_index = idx
                break

        if column_index is None:
            print(f"Column '{column_name}' not found.")
            return

        new_wb = openpyxl.Workbook()
        new_sheet = new_wb.active

        # Copy header row to the new sheet
        for col in range(1, len(header_row) + 1):
            new_sheet.cell(row=1, column=col, value=header_row[col - 1].value)

        filtered_row_count = 0
        # Loop through the rows in the column (skip the first row which is the header)
        for row in sheet.iter_rows(min_row=2, min_col=1, max_col=len(header_row)):
            cell = row[column_index-1]  # Get the cell in the column

            # Check if the cell matches the filter value and if it's highlighted
            if cell.value == filter_value:
                filtered_row_count += 1
                for col_idx, col_cell in enumerate(row, 1):
                    new_cell = new_sheet.cell(row=filtered_row_count + 1, column=col_idx, value=col_cell.value)
                    # Preserve highlighting (if any)
                    if col_cell.fill and col_cell.fill.start_color.index != '00000000' or col_cell.fill != "FFCCCB":
                        new_cell.fill = PatternFill(start_color=col_cell.fill.start_color.index,
                                                end_color=col_cell.fill.end_color.index,
                                                fill_type=col_cell.fill.fill_type
                                                )
        # Save the new workbook with filtered data
        row_counts_by_file={}
        if filtered_row_count > 0:
            new_wb.save(output_file_path)
            print(f"Filtered data saved to {output_file_path}")
            #-------------------------------------------------------------------
            df_file_row_count = pd.read_excel(output_file_path)
            filename = output_file_path.split("/")[-1]
            row_counts_by_file[filename]=len(df_file_row_count)

        else:
            print(f"No rows found with '{column_name}' equal to '{filter_value}'.")

        temp_file_path = os.path.abspath(output_file_path)
        return filtered_data, temp_file_path, row_counts_by_file

    def get_count_totaldata(self):
        #Get the total data of the report
        df_pipeline = pd.read_excel(self.reportpath)
        return len(df_pipeline)

    def count_incorrects(self, file_report:pd.DataFrame):
        # Count the sum of values in each column and return as a dictionary
        column_sum = file_report.apply(lambda x: x.sum()).to_dict()
        return column_sum

    def find_by_files(self):
        # verifying the report by filename for filelevel accuracy
        filenames= self.get_file_names()
        total_file_name = len(filenames)
        if not os.path.exists("temp"):
            os.makedirs("temp")
        for file in filenames:

            outpath = f'temp/{file}_temp.xlsx'
            _, filtereddatapath, data_length_by_file = self.filter_by_category("File_Name", file, outpath)
            res = self.count_column_highlights_byfile(filtereddatapath)

            self.counts[file]=res
        df_result=pd.DataFrame(self.counts).T
        return total_file_name, df_result, data_length_by_file

    def in_pipeline_not_in_gt(self):
        # Get the length of the missing rows in the report
        df_missingrows=pd.read_excel(self.reportpath, sheet_name="InPipelineNotIn_GT")
        total_missing_rows = len(df_missingrows)
        return total_missing_rows

    def extra_rows_in_gt(self):
        df_extragtrows=pd.read_excel(self.reportpath, sheet_name="ExtraRowsinGT")
        df_extragtrows = len(df_extragtrows)
        return df_extragtrows

    def generate_report(self):

        # def analyze_summary_report():

        # Generate a file level and Category level report
        # report = File_Report(report)
        filecounts, df_results, data_count_by_files = self.find_by_files()
        data_count_by_files = self.get_length_by_filename()

        # get the total rows of pipeline sheet
        totalrows = self.get_count_totaldata()
        out = self.count_incorrects(df_results)
        accuracy ={}

        #get total rows of InPipelineNotIn_GT sheet and ExtraRowsinGT sheet
        missingrows = self.in_pipeline_not_in_gt()
        extrarowsingt = self.extra_rows_in_gt()

        # calculate the accuracy of fields
        for field, count in out.items():
            accuracy[field]=round((((totalrows-count)/totalrows)*100),2)

        # Check number of files affected
        files_affected = df_results != 0
        files_affected_counts = files_affected.sum()

        df_category = pd.DataFrame(
            columns=["Iteration Number", "Issue Type", "Issue Level", "Overall Accuracy Percentage",
                    "Number of Files affected"])
        measures = list(df_results.columns)
        for measure in measures:
            category={
                "Iteration Number": "0.0",
                "Issue Type" : measure,
                "Issue Level" : "Field",
                "Overall Accuracy Percentage": float(accuracy[measure]),
                "Number of Files affected": float(files_affected_counts[measure])
            }

            df_category.loc[measures.index(measure)+1] = category.values()


        # droping the key generated column and adding the extra rows of total rows in file
        df_results.drop(columns=['Pseudo_column'], inplace=True)
        df_results["TotalRows"]=df_results.index.map(data_count_by_files)
        columns = ['TotalRows'] + [col for col in df_results.columns if col != 'TotalRows']
        df_results = df_results[columns]

        # Writing the report to the excel file
        output_file = f'{REPORTPATH}/summary_report_{CATEGORY_NAME}.xlsx'
        with pd.ExcelWriter(output_file) as writer:
            df_category.to_excel(writer, sheet_name="Category Level", index=1)
            df_results.to_excel(writer, sheet_name='File Level Accuracy', index=1)
            print(f"Summary report Generated here : {os.path.abspath(output_file)}")

        #wtite the extra data for Data engineers requirement
        wb = openpyxl.load_workbook(output_file)
        file_level_accuracy_report = wb['File Level Accuracy']

        # Writing the total rows of the compared report and insights about the report
        extradata =[[],
                    ["Total data in the pipeline report", totalrows],
                    ["Data in pipeline and not in GT", missingrows],
                    ["Data in Ground Truth not compared with pipeline", extrarowsingt],
                    [],
                    ["ABOUT THE REPORT"],
                    ["Summary reports calculated based only on the highlighted 'Pipeline_Comparission_report' file"],
                    ["Data Comparision has happned only based on the 'pseudo_key' generated"],
                    ["Data  from pipeline file which are Unable to compare or in the 'InPipelinenotInGT' sheet"],
                    ["Data  from Groundtruth file which are Unable to compare with pipeline or in the 'Extrarows in GT' sheet"]]


        for row in extradata:
            file_level_accuracy_report.append(row)

        wb.save(output_file)
        #cleanup directory
        shutil.rmtree("temp")
        print(f"Data written successfully to {output_file}")


# if you want only summary report please uncomment below lines and run this file
# copy the highlighted report
# report ="Reports/highlighted_report_pipelineValidationData_result_temp20250222_234557.xlsx"
# print("Generating the Summary report")
# summary_report = File_Report(report)
# summary_report.generate_report()
# print("Done")
