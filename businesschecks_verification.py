import os

import openpyxl
import pandas as pd
from config import FILE_TO_CHECK_BUSINESS_LOGIC
from openpyxl.styles import PatternFill


class ValidationReportGenerator:
    def __init__(self, mapping_df, lookup_df=None, supplier_category_df=None):
        self.mapping_df = mapping_df
        self.lookup_df = lookup_df
        self.supplier_category_df = supplier_category_df
        self.temp = {}

    def missing_percentage(self, df, mandatory_columns):

        missing_percentage_dfs = []
        for col in mandatory_columns:
            for file in df["File_Name"].drop_duplicates():
                missing_percentage = (
                    df[df["File_Name"] == file][col].isnull().sum()
                    / len(df[df["File_Name"] == file])
                ) * 100
                if missing_percentage > 0:
                    temp_df = pd.DataFrame(
                        {
                            "Category": ["Mandatory Fields"],
                            "Sub Category": ["Missing Mandatory Fields"],
                            "Field": [col],
                            "File Name": [file],
                            "Missing(%)": [missing_percentage],
                        }
                    )
                    missing_percentage_dfs.append(temp_df)
        miss_percent_df = pd.concat(missing_percentage_dfs)
        return miss_percent_df

    def all_null_and_dtype_mismatch(self, df):

        df_dtype = df.dtypes.reset_index()
        df_dtype[0] = df_dtype[0].astype(str)
        df_dtype.rename(columns={"index": "Columns", 0: "dtype"}, inplace=True)
        temp_dfs = []
        for col in df.columns:
            temp_df = None
            if df[col].isnull().all():

                temp_df = pd.DataFrame(
                    {
                        "Category": ["All Null Values"],
                        "Field": [col],
                    }
                )
            elif col in self.mapping_df["Flat_file_columns"].to_list() and (
                df[col].notnull().any()
            ):

                if (
                    df_dtype.loc[df_dtype["Columns"] == col, "dtype"].values[0]
                    != self.mapping_df.loc[
                        self.mapping_df["Flat_file_columns"] == col, "dtype"
                    ].values[0]
                ):

                    temp_df = pd.DataFrame(
                        {
                            "Category": ["Mismatched Data Types"],
                            "Field": [col],
                        }
                    )
            else:
                continue
            temp_dfs.append(temp_df)

        flagged_columns_df = pd.concat(temp_dfs).sort_values(
            by=["Category"], ascending=False
        )

        return flagged_columns_df

    def validation_summary(self, df):
        """
        We are validating the columns in this function.
        Checking the mandatory columns, expected columns, critical columns
        and getting data distribution like null value distribution for the
        mandatory columns.
        expected_columns - Total number of columns which exist in lookup file,
        mandatory_columns - List of Mandatory columns from the lookup file,
        critical_fiels - The columns which are not mandatory columns,
        expected_dtypes - Specifying the data types for  each columns from the
        lookup file

        Args:
            df : Output sheet
        Returns :
            Validation Summary Dataframe

        """

        self.temp["column_indices"] = {
            column: df.columns.get_loc(column) for column in df.columns
        }

        missing_columns = [
            col
            for col in self.mapping_df["Flat_file_columns"].to_list()
            if col not in df.columns
        ]
        extra_columns = [
            col
            for col in df.columns.drop(
                ["Price_range_flag", "Normalized_name_search_flag"]
            )
            if col not in self.mapping_df["Flat_file_columns"].to_list()
        ]

        summary_df = pd.DataFrame(
            columns=[
                "Category",
                "Sub Category",
                "Field",
                "File Name",
                "Missing(%)",
            ]
        )
        summary_df.loc[0] = (
            "Column Structure",
            "Missing Columns",
            None,
            None,
            missing_columns,
        )
        summary_df.loc[1] = (
            "Column Structure",
            "Extra Columns",
            None,
            None,
            extra_columns,
        )

        mandatory_columns = [
            col
            for col in self.mapping_df["Flat_file_columns"]
            if self.mapping_df.loc[
                self.mapping_df["Flat_file_columns"] == col, "Mandatory_column"
            ].values[0]
            == "Y"
        ]
        self.temp["mandatory_columns"] = mandatory_columns
        null_percent_df = self.missing_percentage(df, mandatory_columns)
        flagged_col_df = self.all_null_and_dtype_mismatch(df)
        validation_summary = pd.concat(
            [summary_df, null_percent_df, flagged_col_df]
        )
        return validation_summary

    def get_flagged_index(self, df):

        # Null values
        column_indices = [
            df.columns.get_loc(col) for col in self.temp["mandatory_columns"]
        ]

        index_list = []

        for col_idx in column_indices:
            self.temp[df.columns[col_idx]] = []
            # Find the rows where the column has a null value
            null_rows = df.iloc[:, col_idx].isnull()

            # Add the row and column indices to the list
            for row_idx in null_rows[null_rows].index:
                index_list.append((row_idx + 1, col_idx))
                self.temp[df.columns[col_idx]].append(row_idx)

        # Unknown Values
        unknown_indices = df.map(
            lambda x: isinstance(x, str) and x.lower() == "unknown"
        )
        for row_i, col_i in zip(*unknown_indices.to_numpy().nonzero()):
            index_list.append((row_i + 1, col_i))

        # Total_Price<0
        if "price_columns" in self.temp["config_info"]:
            price_columns = self.temp["config_info"]["price_columns"]
        else:
            price_columns = ["Total_Price"]
        price_columns_idx = [
            df.columns.get_loc(column) for column in price_columns
        ]

        for col_i in price_columns_idx:
            price_indices = (df.iloc[:, col_i] < 0).to_numpy().nonzero()[0]
            for row_i in price_indices:
                index_list.append((row_i + 1, col_i))

        # Price_Date>=2017-01-01
        date_column_idx = self.temp["column_indices"]["Price_Date"]
        date_indices = (
            (df.iloc[:, date_column_idx] < "2017-01-01")
            .to_numpy()
            .nonzero()[0]
        )
        for row_i in date_indices:
            index_list.append((row_i + 1, date_column_idx))

        # Payment_term
        payment_term_column_idx = self.temp["column_indices"]["Payment_Term"]
        payment_term_row_idxs = df[
            ~df["Payment_Term"].str.match(r"^NET\d+$", na=False)
        ].index.to_list()
        for row_idx in payment_term_row_idxs:
            index_list.append((row_idx + 1, payment_term_column_idx))

        return index_list

    def validation_from_lookup(self, df):

        if not isinstance(self.lookup_df, pd.DataFrame):
            print("Invalid Lookup Dataframe. Aborting Lookup Validation...")
            return []

        else:
            index_dict = {}
            for idx, column in enumerate(self.mapping_df["Flat_file_columns"]):
                try:
                    if column == "Product_Service_SKU_Name_Normalized":
                        mapping_values = set(self.lookup_df["normalized_name"])
                        df_values = df[column]
                        index_dict[column] = df_values[
                            ~df_values.isin(mapping_values)
                        ].index

                        self.temp[column] = (
                            self.temp[column] + index_dict[column].to_list()
                        )
                    elif column == "Level 5 Category":
                        mapping_values = set(self.lookup_df["level_5"])
                        df_values = df[column]
                        index_dict[column] = df_values[
                            ~df_values.isin(mapping_values)
                        ].index

                    elif column == "UOM":
                        mapping_values = set(lookup_df["normalized_uom"])
                        df_values = df[column]
                        index_dict[column] = df_values[
                            ~df_values.isin(mapping_values)
                        ].index

                    elif column == "Supplier_Name_Normalized":
                        mapping_values = set(
                            self.supplier_category_df[
                                "supplier_normalized_names"
                            ]
                        )
                        df_values = df[column]
                        index_dict[column] = df_values[
                            ~df_values.isin(mapping_values)
                        ].index

                    elif column == "Product_Service_SKU_Name_Original":
                        mapping_values = set(
                            self.lookup_df[
                                self.lookup_df["remove_flag"]
                                .str.lower()
                                .str.strip()
                                == "yes"
                            ]["original_name"]
                        )
                        df_values = df[column]
                        index_dict[column] = df_values[
                            df_values.isin(mapping_values)
                        ].index

                    elif column == "Unit Price":
                        price_df = self.lookup_df[
                            self.lookup_df["price_outlier"]
                            .str.lower()
                            .str.strip()
                            == "no"
                        ][["normalized_name", "price"]]
                        price_df = price_df.groupby(["normalized_name"]).agg(
                            {"price": ["min", "max"]}
                        )
                        price_df.columns = price_df.columns.get_level_values(1)
                        price_data = df[
                            [
                                "Product_Service_SKU_Name_Normalized",
                                "Unit Price",
                            ]
                        ]
                        combined_df = price_data.merge(
                            price_df,
                            how="left",
                            left_on="Product_Service_SKU_Name_Normalized",
                            right_on="normalized_name",
                        )
                        index_dict[column] = combined_df[
                            (combined_df["Unit Price"] > combined_df["max"])
                            | (combined_df["Unit Price"] < combined_df["min"])
                        ].index
                        self.temp[column] = (
                            self.temp[column] + index_dict[column].to_list()
                        )

                    elif pd.notna(self.mapping_df.loc[idx, "Possible_values"]):
                        mapping_values = (
                            self.mapping_df.loc[idx, "Possible_values"]
                        ).split(",")
                        df_values = df[column]
                        index_dict[column] = df_values[
                            ~df_values.isin(mapping_values)
                        ].index

                except Exception as e:
                    print(
                        "Following error occured in processing column"
                        f" '{column}':\n {repr(e)}"
                    )
                    continue

            index_list = [
                (value + 1, self.temp["column_indices"][key])
                for key, values in index_dict.items()
                for value in values
            ]

            return index_list

    def flagged_cells(self, excel_file_path, index_list, color_code="FFFF00"):

        wb = openpyxl.load_workbook(excel_file_path)

        ws = wb.active
        ws.title = "data"

        fill_color = PatternFill(
            start_color=color_code, end_color=color_code, fill_type="solid"
        )

        # Apply color to the specified cells
        for row_idx, col_idx in index_list:
            cell = ws.cell(row=row_idx + 1, column=col_idx + 1)
            cell.fill = fill_color

        # Save the workbook with the changes
        colored_file_path = "validation_report.xlsx"
        wb.save(colored_file_path)
        return colored_file_path

    def generate_report(self, df, config_info={}):

        self.temp["config_info"] = config_info

        summary_df = self.validation_summary(df)
        flagged_indices = self.get_flagged_index(df)
        indices = flagged_indices + self.validation_from_lookup(df)

        df.loc[:, "Normalized_name_search_flag"] = [
            (
                "Not found"
                if idx in self.temp["Product_Service_SKU_Name_Normalized"]
                else "Found"
            )
            for idx in df.index
        ]
        df.loc[:, "Price_range_flag"] = [
            (
                "Out of Range"
                if (
                    idx in self.temp["Unit Price"]
                    or df.loc[idx, "Normalized_name_search_flag"]
                    == "Not found"
                )
                else "Within Range"
            )
            for idx in df.index
        ]
        excel_file_path = "temp.xlsx"
        df.to_excel(excel_file_path, index=False)

        file_path = self.flagged_cells(excel_file_path, indices)
        print(f"Reports generated here {os.path.abspath(file_path)}")

        with pd.ExcelWriter(file_path, engine="openpyxl", mode="a") as writer:

            summary_df.to_excel(
                writer, sheet_name="validation_summary", index=False
            )

if __name__ == "__main__":


    path = FILE_TO_CHECK_BUSINESS_LOGIC
    lookup_file = f"Test_Data/validation_lookupfile.xlsx"
    df = pd.read_excel(path)
    mapping_df = pd.read_excel(lookup_file, sheet_name="mapping_sheet")
    lookup_df = pd.read_excel(lookup_file, sheet_name="Normalization_Lookup - TBS")
    supplier_category_df = pd.read_excel(
        lookup_file, sheet_name="Supplier to category mapping"
    )

    # if any of the above files are in .csv format,
    # please us pd.read_csv(filename)
    # ex. df = pd.read_csv("tbs_flat_file.csv")

    validator = ValidationReportGenerator(
        mapping_df, lookup_df, supplier_category_df
    )

    validator.generate_report(
        df,
        config_info={"price_columns": ["Total_Price", "Unit Price"]},
    )

